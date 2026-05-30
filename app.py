from flask import Flask , render_template , request
from openpyxl import Workbook ,load_workbook
import os


file_name = "users.xlsx"


def save_excel(user,password,age):
    if not os.path.exists(file_name):
        page_excel = Workbook()
        page =page_excel.active
        page.append(["user","password","age"])
        page_excel.save(file_name)

    page_excel =load_workbook(file_name)
    page=page_excel.active

    page.append([user,password,age])
    page_excel.save(file_name)
    

app = Flask(__name__)

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/registed",methods=["POST"])
def registed():
    name = request.form.get("user")
    password = request.form.get("password")
    age = request.form.get("age")

    if not request.form.get("user") or not request.form.get("password")or  not request.form.get("age"):
        return render_template("not_registed.html")
    
    save_excel(name,password,age)
    return render_template("registed.html")

app.run(debug=True,use_reloader=False)